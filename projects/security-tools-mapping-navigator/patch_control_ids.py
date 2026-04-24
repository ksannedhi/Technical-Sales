"""
Patches the Tool Inventory tab of the template:
  - Updates the hint row placeholder for current_control_id to show
    NIST-only / CIS-only / both formats.
  - Updates the 8 example data rows with a deliberate spread of
    NIST-only, CIS-only, and dual (semicolon-separated) values.
"""

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

TEMPLATE = (
    r"C:\Users\ksann\Downloads\security-tools-mapping-navigator"
    r"\frontend\public\Security_Tools_Mapping_Template.xlsx"
)

wb = openpyxl.load_workbook(TEMPLATE)
ws = wb["Tool Inventory"]

# ── Locate the current_control_id column dynamically ─────────────────────────
header_row = 1
ctrl_id_col = None
for col in range(1, ws.max_column + 1):
    v = ws.cell(row=header_row, column=col).value or ""
    if "current_control_id" in v.lower().replace(" ", "_"):
        ctrl_id_col = col
        break

if ctrl_id_col is None:
    # Try stripping asterisks / spaces
    for col in range(1, ws.max_column + 1):
        v = (ws.cell(row=header_row, column=col).value or "").strip().rstrip("* ").lower()
        if v == "current_control_id":
            ctrl_id_col = col
            break

if ctrl_id_col is None:
    raise RuntimeError("Could not find current_control_id column in Tool Inventory.")

print(f"current_control_id is column {ctrl_id_col} ({get_column_letter(ctrl_id_col)})")

# ── Row 2: hint/placeholder — update to show all three formats ───────────────
hint_cell = ws.cell(row=2, column=ctrl_id_col)
hint_cell.value = (
    "e.g. PR.AA  (NIST only)  |  CIS-5  (CIS only)  |  PR.AA;CIS-5  (both)  "
    "— see reference tables in Instructions"
)

# ── Data rows 4-11: patch current_control_id values ──────────────────────────
# Row layout (based on template build order):
#   Row 4  = MAP-1  Identity / Ping Identity
#   Row 5  = MAP-2  Endpoint / Defender for Endpoint
#   Row 6  = MAP-3  Network  / FortiGate
#   Row 7  = MAP-4  Data     / Mimecast
#   Row 8  = MAP-5  Cloud    / Orca Security
#   Row 9  = MAP-6  AppSec   / F5 WAF
#   Row 10 = MAP-7  SOC      / Splunk
#   Row 11 = MAP-8  Data     / Varonis DSPM

# Deliberate spread:
#   NIST only  → rows 4, 7
#   CIS only   → rows 8, 11
#   Both       → rows 5, 6, 9, 10
CONTROL_IDS = {
    4:  "PR.AA",           # NIST only  — Identity
    5:  "PR.PS;CIS-10",    # Both       — Endpoint (Platform Security + Malware Defenses)
    6:  "PR.IR;CIS-12",    # Both       — Network (Infra Resilience + Network Infra Mgmt)
    7:  "CIS-9",           # CIS only   — Email (Email & Browser Protections)
    8:  "CIS-4",           # CIS only   — Cloud (Secure Configuration)
    9:  "PR.DS;CIS-16",    # Both       — AppSec (Data Security + App Software Security)
    10: "DE.CM;CIS-8",     # Both       — SOC (Continuous Monitoring + Audit Log Mgmt)
    11: "CIS-3;PR.DS",     # Both       — DSPM (Data Protection, CIS first for variety)
}

for data_row, value in CONTROL_IDS.items():
    c = ws.cell(row=data_row, column=ctrl_id_col)
    c.value = value
    c.font = Font(name="Segoe UI", size=10)
    c.alignment = Alignment(horizontal="left", vertical="top")

wb.save(TEMPLATE)
print("Patch applied. Template saved.")
