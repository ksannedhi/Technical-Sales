"""
Adds two new sheets to Security_Tools_Mapping_Template.xlsx:
  1. Discovery Questions  – customer interview guide by domain
  2. Tool Objectives Library – pre-written control_objective text for 30 tool categories

Inserts both sheets after the Instructions tab so the logical flow is:
  Instructions → Discovery Questions → Tool Objectives Library → Tool Inventory
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEMPLATE = (
    r"C:\Users\ksann\Downloads\security-tools-mapping-navigator"
    r"\frontend\public\Security_Tools_Mapping_Template.xlsx"
)

# ── Palette ──────────────────────────────────────────────────────────────────
CA   = "0F8B6B"   # accent green
CD   = "114B5F"   # dark teal
CB   = "E6EEEA"   # light mint bg
CW   = "FFFFFF"
CT   = "152222"
CM   = "5B6F6B"
CBR  = "CFE0DA"

DOMAIN_BG = {
    "Identity": "EAF0FF", "Endpoint": "E8F5E9",
    "Network":  "E3F2FD", "Data":     "FFF8E1",
    "Cloud":    "F3E5F5", "AppSec":   "FCEAE8",
    "SOC":      "E0F7F4",
}
DOMAIN_FG = {
    "Identity": "283593", "Endpoint": "1B5E20",
    "Network":  "0D47A1", "Data":     "E65100",
    "Cloud":    "6A1B9A", "AppSec":   "B71C1C",
    "SOC":      "004D40",
}

# ── Style helpers ─────────────────────────────────────────────────────────────
def _b():
    s = Side(style="thin", color=CBR)
    return Border(left=s, right=s, top=s, bottom=s)

def _bot():
    s = Side(style="thin", color=CBR)
    return Border(bottom=s)

def sc(cell, bold=False, fg=CT, bg=None, size=10,
       wrap=False, ha="left", va="top", bdr=None):
    cell.font = Font(name="Segoe UI", bold=bold, color=fg, size=size)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if bdr:
        cell.border = bdr

def sheet_title(ws, text, sub, cols):
    last = get_column_letter(cols)
    ws.merge_cells(f"A1:{last}1")
    ws["A1"] = text
    sc(ws["A1"], bold=True, fg=CW, bg=CD, size=13, ha="center", va="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{last}2")
    ws["A2"] = sub
    sc(ws["A2"], fg=CM, bg=CB, size=9, wrap=True, ha="left", va="center")
    ws.row_dimensions[2].height = 48

    ws.row_dimensions[3].height = 6   # visual spacer

def col_headers(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=row, column=i, value=h)
        sc(c, bold=True, fg=CW, bg=CA, size=10, bdr=_b())
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 22

def domain_hdr(ws, row, domain, cols):
    last = get_column_letter(cols)
    ws.merge_cells(f"A{row}:{last}{row}")
    c = ws[f"A{row}"]
    c.value = f"  {domain}"
    sc(c, bold=True, fg=DOMAIN_FG.get(domain, CT),
       bg=DOMAIN_BG.get(domain, "F5F5F5"), size=10, va="center")
    ws.row_dimensions[row].height = 18

def data_row(ws, row, values, h=60, wrap_cols=None):
    wrap_cols = wrap_cols or set()
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        sc(c, wrap=(i in wrap_cols), va="top", bdr=_bot())
    ws.row_dimensions[row].height = h


# ═══════════════════════════════════════════════════════════════════════════════
#  DISCOVERY QUESTIONS DATA
# ═══════════════════════════════════════════════════════════════════════════════
DISCOVERY_HEADERS = [
    "Domain", "#",
    "Discovery Question",
    "What a Good Answer Looks Like",
    "Red Flags to Watch For",
]
DISCOVERY_WIDTHS = [12, 4, 46, 42, 42]

DISCOVERY_DATA = [
    ("Identity", [
        ("1",
         "What IAM or SSO solution is in use today? Is it cloud-native or on-premises?",
         "Named vendor enforced for all apps — Okta, Entra ID, Ping Identity. "
         "Cloud-native SSO preferred over on-prem AD federation.",
         "No SSO — each application has its own login. On-prem Active Directory only, "
         "no federation to cloud apps."),
        ("2",
         "Is MFA enforced for all users, including admins and remote access?",
         "100% MFA coverage with phishing-resistant methods (FIDO2 / Passkeys) for admins. "
         "No legacy auth protocols (NTLM, Basic Auth) enabled.",
         "MFA optional or enforced only on VPN. SMS OTP is the only MFA method. "
         "Admin accounts are exempt from MFA policy."),
        ("3",
         "How are privileged and admin accounts managed? Is there a PAM solution?",
         "CyberArk, BeyondTrust, or Delinea deployed with session recording, password vaulting, "
         "and just-in-time access grants.",
         "Shared admin passwords stored in a spreadsheet. No audit trail for privileged sessions. "
         "Domain Admin used for everyday tasks."),
        ("4",
         "Is there an Identity Governance (IGA) platform for access reviews and certification?",
         "SailPoint or Saviynt running quarterly access certifications. "
         "Joiner / Mover / Leaver lifecycle is automated.",
         "No access reviews. Off-boarding leaves accounts active for weeks or longer. "
         "No inventory of what entitlements each user holds."),
        ("5",
         "How are service accounts and API credentials managed and rotated?",
         "Secrets vault (HashiCorp Vault, CyberArk) with automated rotation. "
         "No long-lived static credentials hard-coded in source code.",
         "Credentials hard-coded in scripts and repositories. "
         "Service account passwords unchanged for years. No inventory of service accounts."),
    ]),
    ("Endpoint", [
        ("1",
         "Which EDR or XDR platform is deployed, and what percentage of endpoints are covered?",
         "CrowdStrike, SentinelOne, or Defender for Endpoint at 95%+ coverage. "
         "Unified console across workstations, servers, and mobile.",
         "Coverage below 80%. Multiple competing agents on the same endpoints. "
         "Contractor and remote devices excluded from EDR."),
        ("2",
         "Are there unmanaged or BYOD devices accessing corporate resources?",
         "NAC controls enforce device health checks before network access is granted. "
         "BYOD limited to email via an MDM-managed container (Intune / Jamf).",
         "BYOD unrestricted on the corporate network. No NAC. "
         "Personal devices have full access to internal file shares and applications."),
        ("3",
         "What is the vulnerability management program? How quickly are critical patches applied?",
         "Tenable or Qualys scanning weekly. Critical CVEs patched within 72 hours. "
         "SLAs tracked and reported in monthly metrics.",
         "No formal VM program. Patches applied ad hoc or only during quarterly windows. "
         "No SLA tracking. Unapplied critical CVEs measured in months."),
        ("4",
         "How are endpoint configurations hardened? Is there a CIS benchmark baseline in use?",
         "CIS Level 1 or Level 2 applied via Microsoft Intune or SCCM. "
         "Configuration drift detection alerts the SOC when deviation is detected.",
         "Default OS configurations with no hardening standard. "
         "No tool to detect when configuration changes post-deployment."),
        ("5",
         "Are servers (Windows and Linux) covered by the same security tooling as workstations?",
         "Unified EDR deployed to both workstations and servers. "
         "Separate detection policies tuned per asset class.",
         "Servers excluded from EDR or running legacy AV only. "
         "Server patching managed by a separate team with no SOC visibility."),
    ]),
    ("Network", [
        ("1",
         "What firewall platform protects the perimeter and east-west traffic?",
         "NGFW (Palo Alto, Fortinet FortiGate, Check Point) with IPS in blocking mode. "
         "Microsegmentation enforced between zones and workloads.",
         "Legacy stateful firewalls at the perimeter only. Flat network — no east-west controls "
         "or microsegmentation. IPS in detection mode only."),
        ("2",
         "Is there a Zero Trust or SASE platform replacing or supplementing legacy VPN?",
         "Zscaler, Netskope, or Prisma Access enforcing continuous user and device trust "
         "before granting application-level access.",
         "All remote access via split-tunnel VPN. No device health or user context evaluated. "
         "VPN grants broad network access on connection."),
        ("3",
         "Is DNS security in place to block command-and-control and malicious domain lookups?",
         "Cisco Umbrella or Infoblox BloxOne with threat intelligence feeds. "
         "DNS-layer blocking active on all endpoints — including off-network.",
         "No DNS filtering. Default ISP or cloud provider DNS in use. "
         "No DNS query logging or visibility for threat hunting."),
        ("4",
         "Is network traffic monitored for lateral movement and anomalies — not just perimeter logs?",
         "NDR solution (Vectra AI, Darktrace, ExtraHop) with east-west traffic baselining "
         "and automated anomaly alerting.",
         "Only perimeter firewall logs reach the SOC. No east-west visibility. "
         "Lateral movement from a compromised host would go undetected."),
    ]),
    ("Data", [
        ("1",
         "What email security solution is in place for inbound and outbound threats?",
         "Mimecast or Proofpoint Advanced with URL rewriting, attachment sandboxing, "
         "impersonation protection, and outbound DLP rules active.",
         "Default Exchange Online Protection only. No sandboxing or advanced BEC / "
         "impersonation controls configured."),
        ("2",
         "Is there a DLP solution monitoring sensitive data in motion, at rest, and in use?",
         "Microsoft Purview or Digital Guardian with active classification-based policies "
         "on email, cloud storage, and endpoints.",
         "No DLP. USB transfers and cloud storage uploads of sensitive data are "
         "unrestricted and unmonitored."),
        ("3",
         "Is sensitive data classified across cloud storage, endpoints, and email?",
         "Automated classification (Purview Sensitivity Labels or Varonis) applied to "
         "new and existing data. Regular policy reviews performed quarterly.",
         "Manual classification only — or none at all. "
         "No way to locate where PII or confidential data actually resides."),
        ("4",
         "Is there a DSPM or data discovery tool providing visibility into cloud data stores?",
         "Varonis, BigID, or Laminar scanning S3 buckets, SharePoint, and databases. "
         "Data risk posture tracked on a weekly basis.",
         "No visibility into sensitive data in cloud environments. "
         "No data store inventory or data ownership mapping."),
    ]),
    ("Cloud", [
        ("1",
         "What cloud providers are in scope (AWS, Azure, GCP, SaaS)? "
         "Is there a central cloud account inventory?",
         "All cloud accounts registered in a central inventory. "
         "CSPM console has 100% estate visibility across all providers.",
         "Shadow IT cloud usage with no central registry. Business units spinning up "
         "accounts outside any security oversight or billing visibility."),
        ("2",
         "Is there a CSPM or CNAPP solution monitoring cloud posture continuously?",
         "Wiz, Prisma Cloud, or Orca Security scanning all cloud workloads continuously. "
         "Policy violations trigger alerts with a defined 24-hour remediation SLA.",
         "Manual cloud security reviews only. No continuous misconfiguration detection. "
         "Last cloud security audit was more than 6 months ago."),
        ("3",
         "How are cloud entitlements and permissions governed? Is least privilege enforced?",
         "CIEM solution (Ermetic, Sonrai Security) reviews effective cloud permissions. "
         "Wildcard IAM policies flagged automatically and remediated within SLA.",
         "Wildcard IAM policies in use across cloud accounts. No visibility into effective "
         "permissions. No regular entitlement reviews — over-privileged identities accumulate."),
        ("4",
         "Are container and Kubernetes workloads scanned for vulnerabilities and runtime threats?",
         "Aqua Security or Sysdig scanning images in CI/CD. Privileged containers blocked "
         "in production. Runtime anomaly detection active on Kubernetes nodes.",
         "No container image scanning. Privileged containers running in production. "
         "No runtime protection on Kubernetes workloads."),
    ]),
    ("AppSec", [
        ("1",
         "Is there a WAF protecting all externally facing web applications and APIs?",
         "F5, Imperva, or Cloudflare WAF in full blocking mode on all external apps. "
         "OWASP Top 10 rule sets active and tuned after each pen test.",
         "WAF in detection-only mode or not covering all applications. "
         "Custom-built internal apps completely unprotected."),
        ("2",
         "Are all APIs inventoried and monitored for abuse and security anomalies?",
         "Salt Security or Noname with a complete API inventory, schema validation, "
         "and behavioural anomaly detection active.",
         "No API security posture management. Shadow and undocumented APIs in production. "
         "No monitoring for API abuse or credential stuffing."),
        ("3",
         "Is DAST or penetration testing performed on applications on a regular schedule?",
         "Quarterly automated DAST (Invicti, Burp Suite Enterprise) plus an annual "
         "third-party pen test. All findings formally tracked to closure.",
         "No regular application testing. Pen tests triggered only by compliance audits. "
         "Findings not formally tracked or assigned owners."),
        ("4",
         "Is SAST or SCA integrated into the CI/CD pipeline with build-breaking policies?",
         "Checkmarx, Veracode, or Snyk embedded in CI/CD pipeline. Critical findings "
         "block deployment. Mean-time-to-remediate tracked in monthly metrics.",
         "No SAST or SCA in the build pipeline. Developers manually check libraries — "
         "or do not check at all. Vulnerable open-source components shipped to production."),
    ]),
    ("SOC", [
        ("1",
         "What SIEM platform is in use, and what is the current log coverage across the estate?",
         "Splunk or Microsoft Sentinel ingesting 95%+ of log sources — endpoints, cloud, "
         "network, identity, and applications. Normalised into common schema.",
         "Log coverage below 70%. OT, cloud, or identity logs not ingested. "
         "SIEM alert backlog growing faster than it is worked by analysts."),
        ("2",
         "Is there a SOAR platform for automated alert triage and incident response?",
         "Palo Alto XSOAR or Splunk SOAR with playbooks covering 50%+ of common incident types. "
         "MTTD and MTTR formally tracked and reviewed weekly.",
         "All incident response is fully manual. No automation for common alert triage. "
         "Analysts spending the majority of their time on repetitive, low-value tasks."),
        ("3",
         "What are the current Mean Time to Detect (MTTD) and Mean Time to Respond (MTTR) metrics?",
         "MTTD under 1 hour and MTTR under 4 hours for high-severity incidents. "
         "Metrics published in weekly SOC review meetings with trend analysis.",
         "No MTTD or MTTR tracking. Incidents sometimes discovered days after initial compromise. "
         "No formal post-incident review or lessons-learned process."),
        ("4",
         "Is there 24x7 SOC coverage — in-house, MDR, or a hybrid model?",
         "24x7 coverage via in-house SOC or a named MDR partner (Arctic Wolf, Expel, Secureworks) "
         "with documented escalation paths and SLA commitments.",
         "Business-hours-only monitoring. No on-call rotation. High-severity incidents "
         "outside working hours go undetected until the next working day."),
    ]),
]


# ═══════════════════════════════════════════════════════════════════════════════
#  TOOL OBJECTIVES LIBRARY DATA
# ═══════════════════════════════════════════════════════════════════════════════
TOL_HEADERS = [
    "Tool Category",
    "Common Vendors / Products",
    "Control Domain",
    "Typical Control Objective  (copy into Tool Inventory)",
    "Framework Alignment",
    "Coverage Level Guidance",
    "Presales Notes",
]
TOL_WIDTHS = [26, 36, 14, 50, 22, 28, 38]

TOL_DATA = [
    ("Identity", [
        ("IAM / Single Sign-On (SSO)",
         "Okta, Microsoft Entra ID, Ping Identity, OneLogin, JumpCloud",
         "Identity",
         "Centralised authentication and single sign-on for all enterprise applications and user accounts",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if all apps federated; Partially Covered if select apps only",
         "Ask: % of apps in SSO, MFA enforcement rate, whether legacy auth protocols are disabled"),
        ("Multi-Factor Authentication (MFA)",
         "Duo Security, Microsoft Authenticator, RSA SecurID, Yubico",
         "Identity",
         "Multi-factor authentication enforcement for all user accounts, remote access, and privileged operations",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if 100% of users enrolled; Partially Covered if partial rollout or SMS-only",
         "Check if MFA is phishing-resistant (FIDO2) for admins. SMS OTP is a partial control only."),
        ("Privileged Access Management (PAM)",
         "CyberArk, BeyondTrust, Delinea (Thycotic), HashiCorp Vault",
         "Identity",
         "Privileged access management and session recording for administrator and service accounts",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if vaulting + recording active; Partially Covered if vaulting only",
         "Key differentiator: just-in-time access vs always-on. Ask about service account coverage too."),
        ("Identity Governance & Administration (IGA)",
         "SailPoint IdentityNow, Saviynt, One Identity, IBM Security Verify",
         "Identity",
         "Identity governance, access certification, and lifecycle management for all enterprise user accounts",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if access reviews automated; Partially Covered if manual reviews only",
         "IGA is distinct from SSO/MFA. Many orgs have SSO but lack governance. Ask about access review frequency."),
        ("Directory Services",
         "Microsoft Active Directory, Azure AD DS, OpenLDAP, Okta LDAP Interface",
         "Identity",
         "Central directory and authentication infrastructure for enterprise user and device identities",
         "NIST-CSF-2.0;CIS-v8.1",
         "Partially Covered — directory alone does not constitute full IAM coverage",
         "Usually present, often poorly secured. Check for AD tiering, stale accounts, and Group Policy hygiene."),
    ]),
    ("Endpoint", [
        ("EDR / XDR",
         "CrowdStrike Falcon, SentinelOne, Microsoft Defender for Endpoint, Cortex XDR",
         "Endpoint",
         "Endpoint detection and response for workstations, servers, and mobile devices",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if 95%+ deployment; Partially Covered if gaps exist",
         "Coverage % is the critical metric. Ask about server coverage separately — often excluded."),
        ("Vulnerability Management",
         "Tenable.io, Qualys VMDR, Rapid7 InsightVM, Microsoft Defender Vulnerability Management",
         "Endpoint",
         "Continuous vulnerability scanning and patch management for all enterprise endpoints and servers",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if scanning + patching SLA in place; Partially Covered if scan-only",
         "Scanning without a patching SLA is Partially Covered at best. Ask about mean time to patch critical CVEs."),
        ("Endpoint Management / MDM",
         "Microsoft Intune, Jamf Pro, VMware Workspace ONE, Kandji",
         "Endpoint",
         "Endpoint configuration management, policy enforcement, and software deployment for managed devices",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if CIS benchmark applied with drift detection; Partially Covered if basic MDM only",
         "MDM != EDR. Both are needed. MDM handles config/compliance; EDR handles threat detection."),
        ("Antivirus / EPP (legacy)",
         "Symantec Endpoint Protection, McAfee/Trellix, Trend Micro, ESET",
         "Endpoint",
         "Malware prevention and endpoint protection for workstations and servers via signature and heuristic detection",
         "NIST-CSF-2.0;CIS-v8.1",
         "Minimally Covered — legacy AV without EDR leaves significant detection gaps",
         "Flag as a gap opportunity. Legacy AV is a consolidation candidate when introducing modern EDR."),
    ]),
    ("Network", [
        ("Next-Generation Firewall (NGFW)",
         "Palo Alto Networks, Fortinet FortiGate, Check Point, Cisco Firepower",
         "Network",
         "Network boundary and segmentation controls for perimeter and east-west traffic inspection",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if IPS in blocking mode and microsegmentation active; Partially Covered otherwise",
         "Ask about east-west coverage — not just perimeter. IPS in detection mode is Partially Covered."),
        ("Network Detection & Response (NDR)",
         "Vectra AI, Darktrace, ExtraHop Reveal(x), Cisco Stealthwatch",
         "Network",
         "Network traffic analysis and anomaly detection for lateral movement, C2, and insider threats",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if baselining + alerting active; Partially Covered if passive capture only",
         "Often absent even in mature orgs. Strong differentiator for SIEM+NDR integration conversations."),
        ("ZTNA / SASE",
         "Zscaler, Netskope, Palo Alto Prisma Access, Cloudflare Access, Cato Networks",
         "Network",
         "Zero Trust network access and secure web gateway for remote users and cloud application access",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if ZTNA replaces VPN with device trust; Partially Covered if coexisting with VPN",
         "Ask if this is a VPN replacement or a parallel solution. Coexistence often creates policy gaps."),
        ("DNS Security",
         "Cisco Umbrella, Infoblox BloxOne, Akamai ETP, Palo Alto DNS Security",
         "Network",
         "DNS-layer threat protection to block malicious domain lookups, phishing, and C2 communications",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if deployed to all endpoints including off-network; Partially Covered if on-network only",
         "Quick win — DNS security is low-cost and high-value. Often missing or limited to corporate network only."),
    ]),
    ("Data", [
        ("Email Security / SEG",
         "Mimecast, Proofpoint, Microsoft Defender for Office 365, Barracuda",
         "Data",
         "Email threat prevention and data security for inbound phishing, BEC, malware, and outbound data loss",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if sandboxing + DLP + impersonation protection all active; Partially Covered otherwise",
         "Distinguish between E3 Defender (basic) and E5 / Proofpoint Advanced (full). DLP rules often unconfigured."),
        ("Data Loss Prevention (DLP)",
         "Microsoft Purview, Digital Guardian, Symantec DLP, Forcepoint DLP",
         "Data",
         "Data loss prevention and policy enforcement for sensitive data in motion, at rest, and in use",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if policy enforcement active on email + endpoint + cloud; Partially Covered if monitoring only",
         "DLP in monitoring mode is common — ask if it blocks or only alerts. Endpoint DLP often unenforced."),
        ("Data Security Posture Management (DSPM)",
         "Varonis, BigID, Laminar, Securiti, Normalyze",
         "Data",
         "Data security posture management and sensitive data discovery across cloud storage and databases",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if all cloud data stores scanned with risk tracking; Partially Covered if partial coverage",
         "Emerging category — many orgs lack this entirely. Strong conversation starter for data sprawl risk."),
        ("Backup & Disaster Recovery",
         "Veeam, Commvault, Zerto, Rubrik, Cohesity",
         "Data",
         "Secure backup, replication, and disaster recovery to ensure business continuity and data resilience",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if backup tested and immutable copies maintained; Partially Covered if untested",
         "Ask when backups were last tested end-to-end. Immutability (air gap / WORM) is key against ransomware."),
    ]),
    ("Cloud", [
        ("CSPM / CNAPP",
         "Wiz, Palo Alto Prisma Cloud, Orca Security, Lacework, Aqua",
         "Cloud",
         "Cloud security posture and workload protection for cloud-hosted infrastructure and applications",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if all cloud accounts scanned with policy enforcement; Partially Covered if partial estate",
         "CNAPP = CSPM + CWPP + CIEM in one platform. Check whether IaC scanning and CI/CD integration are active."),
        ("Cloud Workload Protection (CWPP)",
         "Microsoft Defender for Cloud, Sysdig, Aqua Security, Lacework",
         "Cloud",
         "Runtime threat detection and workload protection for cloud VMs, containers, and serverless functions",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if agent deployed to all workloads with runtime protection; Partially Covered otherwise",
         "Often bundled in Defender for Cloud or CNAPP platforms. Ask about serverless and container coverage."),
        ("Cloud Infrastructure Entitlement Mgmt (CIEM)",
         "Ermetic (now Tenable), Sonrai Security, CrowdStrike Horizon, Wiz CIEM",
         "Cloud",
         "Cloud entitlement governance and least-privilege enforcement across multi-cloud IAM roles and permissions",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if effective-permission analysis with remediation; Partially Covered if discovery only",
         "Most orgs are over-permissioned in cloud. CIEM identifies unused and excessive permissions at scale."),
        ("Container Security",
         "Aqua Security, Sysdig Secure, Snyk Container, Trivy (open-source)",
         "Cloud",
         "Container image scanning and Kubernetes security for CI/CD pipelines and production workloads",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if image scanning + runtime protection active; Partially Covered if CI/CD only",
         "Ask if scanning happens at build time (CI/CD) and also at runtime. Runtime is where most orgs have gaps."),
    ]),
    ("AppSec", [
        ("Web Application Firewall (WAF)",
         "F5 Advanced WAF, Imperva Cloud WAF, Cloudflare WAF, AWS WAF, Azure WAF",
         "AppSec",
         "Web application and API protection against OWASP Top 10 threats, injection attacks, and automated bots",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if blocking mode on all apps; Partially Covered if detection-only or partial app coverage",
         "Key question: is it in blocking or detection mode? Detection-only WAF is Minimally Covered at best."),
        ("API Security",
         "Salt Security, Noname Security, Traceable AI, Cequence Security",
         "AppSec",
         "API discovery, runtime protection, and abuse prevention for internal and external API endpoints",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if full API inventory + anomaly detection + enforcement; Partially Covered if discovery only",
         "Most WAFs do not provide full API security. Shadow API discovery is the key capability to highlight."),
        ("Dynamic Application Security Testing (DAST)",
         "Invicti (Netsparker), HCL AppScan, Burp Suite Enterprise, Rapid7 InsightAppSec",
         "AppSec",
         "Dynamic security testing and vulnerability discovery for web applications in staging and production",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if automated quarterly scans + annual pen test; Partially Covered if annual pen test only",
         "DAST tests running apps. Ask whether it runs in CI/CD or only pre-release. Annual pen tests alone are Partial."),
        ("SAST / Code Scanning",
         "Checkmarx, Veracode, Semgrep, Fortify, SonarQube",
         "AppSec",
         "Static application security testing integrated into the CI/CD pipeline for vulnerability detection in source code",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if CI/CD integrated with build gates; Partially Covered if scan-only without enforcement",
         "Scan without blocking = Partial. Ask if critical findings break the build. IDE plugins alone are Minimal."),
        ("Software Composition Analysis (SCA)",
         "Snyk, Black Duck, Mend (WhiteSource), FOSSA, Dependabot",
         "AppSec",
         "Open-source dependency and software composition analysis for vulnerable third-party libraries in CI/CD",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if CI/CD integrated with license + vulnerability policies; Partially Covered if ad hoc scans",
         "Ask about SBOM generation. Regulatory pressure (Executive Order 14028) is driving SBOM requirements."),
    ]),
    ("SOC", [
        ("SIEM",
         "Splunk Enterprise Security, Microsoft Sentinel, IBM QRadar, Elastic Security, LogRhythm",
         "SOC",
         "Centralised security monitoring, log management, and threat detection across the hybrid environment",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if 90%+ log coverage + tuned use cases; Partially Covered if log gaps or alert backlog",
         "Log coverage % is the key metric. Ask about OT, cloud, and identity log ingestion specifically."),
        ("SOAR",
         "Palo Alto XSOAR, Splunk SOAR, ServiceNow SecOps, Microsoft Sentinel Automation",
         "SOC",
         "Security orchestration, automation, and response for accelerated incident triage and remediation",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if 50%+ of incidents on automated playbooks; Partially Covered if limited automation",
         "SOAR without playbook coverage is a tool, not a capability. Ask: what % of alerts are auto-triaged?"),
        ("Threat Intelligence Platform (TIP)",
         "Recorded Future, ThreatConnect, MISP (open-source), Mandiant Advantage",
         "SOC",
         "Threat intelligence ingestion, enrichment, and operationalisation for proactive detection and response",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if TI feeds auto-enrich SIEM alerts; Partially Covered if manual lookup only",
         "Many orgs have TI feeds but no automation. Operationalised TI = indicators auto-blocked in NGFW + SIEM."),
        ("MDR / MSSP",
         "Arctic Wolf, Expel, Secureworks Taegis, Palo Alto Unit 42, CrowdStrike Falcon Complete",
         "SOC",
         "Managed detection and response providing 24x7 threat monitoring, triage, and incident response",
         "NIST-CSF-2.0;CIS-v8.1",
         "Fully Covered if 24x7 MDR with SLA-backed response; Partially Covered if monitoring-only MSSP",
         "Distinguish MDR (detect + respond) from MSSP (monitor + alert). MDR is a stronger control posture."),
    ]),
]


# ═══════════════════════════════════════════════════════════════════════════════
#  BUILD SHEETS
# ═══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.load_workbook(TEMPLATE)
existing = wb.sheetnames

# Remove old versions if re-running
for name in ["Discovery Questions", "Tool Objectives Library"]:
    if name in existing:
        del wb[name]

# ── Discovery Questions ───────────────────────────────────────────────────────
ws_dq = wb.create_sheet("Discovery Questions", 1)

sheet_title(
    ws_dq,
    "Security Discovery Questions",
    "Use these questions during a customer discovery session to understand their current security tool landscape. "
    "Record the tools the customer mentions directly into the Tool Inventory tab — one row per tool or control area.",
    len(DISCOVERY_HEADERS),
)
col_headers(ws_dq, 4, DISCOVERY_HEADERS, DISCOVERY_WIDTHS)

row = 5
for domain, questions in DISCOVERY_DATA:
    domain_hdr(ws_dq, row, domain, len(DISCOVERY_HEADERS))
    row += 1
    for num, question, good, red in questions:
        data_row(ws_dq, row,
                 [domain, num, question, good, red],
                 h=72,
                 wrap_cols={3, 4, 5})
        ws_dq.cell(row=row, column=1).alignment = Alignment(
            horizontal="left", vertical="top")
        row += 1

ws_dq.freeze_panes = "C5"

# ── Tool Objectives Library ───────────────────────────────────────────────────
ws_tol = wb.create_sheet("Tool Objectives Library", 2)

sheet_title(
    ws_tol,
    "Tool Objectives Library",
    "Pre-written control_objective text for 30 common tool categories. "
    "Copy the 'Typical Control Objective' into the Tool Inventory when filling in a customer's stack. "
    "Adjust wording to match what the customer has actually deployed.",
    len(TOL_HEADERS),
)
col_headers(ws_tol, 4, TOL_HEADERS, TOL_WIDTHS)

row = 5
for domain, tools in TOL_DATA:
    domain_hdr(ws_tol, row, domain, len(TOL_HEADERS))
    row += 1
    for tool_cat, vendors, domain_val, objective, fw, coverage, notes in tools:
        data_row(ws_tol, row,
                 [tool_cat, vendors, domain_val, objective, fw, coverage, notes],
                 h=54,
                 wrap_cols={1, 2, 4, 6, 7})
        row += 1

ws_tol.freeze_panes = "D5"

# ── Reorder: Instructions → Discovery → Tool Objectives → Tool Inventory ─────
sheet_order = ["Instructions", "Discovery Questions", "Tool Objectives Library"]
for name in wb.sheetnames:
    if name not in sheet_order:
        sheet_order.append(name)

for idx, name in enumerate(sheet_order):
    if name in wb.sheetnames:
        wb.move_sheet(name, offset=wb.sheetnames.index(name) - idx)

wb.save(TEMPLATE)
print("Template saved OK:", TEMPLATE)
